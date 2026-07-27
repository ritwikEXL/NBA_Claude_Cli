"""
Generate CareIntel demo Excel workbook for Data Sources upload testing.
Produces: careintel_demo_data.xlsx  (3 sheets: Members, Gaps, Plans)
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import date, timedelta
import random, os

random.seed(42)

# ── Palette ───────────────────────────────────────────────────────────────────
ORANGE     = "FB4E0B"
ORANGE_L   = "FFF1EB"
HEADER_FG  = "FFFFFF"
ALT_ROW    = "FEF5F1"
BORDER_CLR = "FAD4C7"

def hdr_cell(ws, row, col, value, width=18):
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(bold=True, color=HEADER_FG, size=10)
    c.fill = PatternFill("solid", fgColor=ORANGE)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Side(style="thin", color=BORDER_CLR)
    c.border = Border(left=thin, right=thin, top=thin, bottom=thin)
    ws.column_dimensions[get_column_letter(col)].width = width
    return c

def data_cell(ws, row, col, value, alt=False, fmt=None):
    c = ws.cell(row=row, column=col, value=value)
    c.fill = PatternFill("solid", fgColor=ALT_ROW if alt else "FFFFFF")
    thin = Side(style="thin", color=BORDER_CLR)
    c.border = Border(left=thin, right=thin, top=thin, bottom=thin)
    c.alignment = Alignment(vertical="center")
    if fmt:
        c.number_format = fmt
    return c

# ── Data definitions ──────────────────────────────────────────────────────────
PLANS = [
    ("P001", "Aetna Medicare Choice PPO (Northeast)",   "Northeast", "MAPD", 3.5, 4.0, 450_000_000, 45_000),
    ("P002", "Aetna Medicare Premier PPO (Southeast)",  "Southeast", "MAPD", 4.0, 4.5, 380_000_000, 38_000),
    ("P003", "Aetna Medicare DSNP Community (Midwest)", "Midwest",   "DSNP", 3.0, 4.0, 280_000_000, 28_000),
    ("P004", "UHC Medicare Advantage Value (West)",     "West",      "MAPD", 4.5, 5.0, 520_000_000, 52_000),
    ("P005", "UHC Medicare Signature PPO (West)",       "West",      "MAPD", 2.5, 3.0, 210_000_000, 21_000),
]

MEASURES = ["BCS", "COL", "EED", "AFV", "CBP"]
LANGUAGES = ["EN"]*60 + ["ES"]*25 + ["ZH"]*10 + ["VI"]*5
LITERACY   = ["High"]*40 + ["Medium"]*35 + ["Low"]*25
SES        = ["High"]*20 + ["Mid"]*50 + ["Low"]*30
CHANNELS   = ["EMAIL", "SMS", "CALL"]
INCENTIVES = ["GIFTCARD_25", "GIFTCARD_15", "FIT_KIT_MAILER", "TRANSPORT_VOUCHER", "NONE"]

N_MEMBERS = 200

# ── Generate members ──────────────────────────────────────────────────────────
members = []
for i in range(1, N_MEMBERS + 1):
    mid = f"MEM{10000 + i:05d}"
    dob = date(1930, 1, 1) + timedelta(days=random.randint(0, 365*35))  # age 55-90
    gender = random.choice(["F","F","M"])
    lang = random.choice(LANGUAGES)
    lit  = random.choice(LITERACY)
    ses  = random.choice(SES)
    # channel prefs
    email_ok = lang == "EN" and lit in ("High","Medium")
    sms_ok   = lit in ("High","Medium") or random.random() > 0.6
    call_ok  = True
    if not email_ok and not sms_ok:
        call_ok = True
    pref = "EMAIL" if email_ok else ("SMS" if sms_ok else "CALL")
    dnc  = random.random() < 0.03
    members.append((mid, dob.strftime("%Y-%m-%d"), gender, lang, lit, ses,
                    str(email_ok).lower(), str(sms_ok).lower(), str(call_ok).lower(),
                    pref, str(dnc).lower()))

# ── Generate gaps ─────────────────────────────────────────────────────────────
gaps = []
gap_open_base = date(2026, 1, 1)
for mem in members:
    mid = mem[0]
    plan = random.choice(PLANS)[0]
    n_gaps = random.randint(1, 3)
    used = set()
    for _ in range(n_gaps):
        mcode = random.choice(MEASURES)
        if mcode in used:
            continue
        used.add(mcode)
        days_open = random.randint(45, 300)
        open_date = (gap_open_base + timedelta(days=random.randint(0, 180))).strftime("%Y-%m-%d")
        status = "Open" if days_open > 90 else "Borderline"
        risk = round(random.uniform(0.3, 0.95), 2)
        prop = round(random.uniform(0.35, 0.92), 2)
        ch = mem[9]  # preferred channel
        inc = random.choice(INCENTIVES)
        gaps.append((mid, mcode, plan, 2026, status, open_date, risk, days_open, prop, ch, inc))

# ─────────────────────────────────────────────────────────────────────────────
wb = openpyxl.Workbook()

# ── Sheet 1: Members ──────────────────────────────────────────────────────────
ws1 = wb.active
ws1.title = "Member Roster"
ws1.row_dimensions[1].height = 28

MEM_COLS = [
    ("member_id", 14), ("date_of_birth", 14), ("gender", 9), ("language_preference", 14),
    ("digital_literacy_segment", 20), ("socioeconomic_segment", 18),
    ("email_allowed", 13), ("sms_allowed", 11), ("call_allowed", 11),
    ("preferred_channel", 16), ("do_not_contact_flag", 18),
]
for ci, (name, w) in enumerate(MEM_COLS, 1):
    hdr_cell(ws1, 1, ci, name, w)

for ri, row in enumerate(members, 2):
    alt = ri % 2 == 0
    for ci, val in enumerate(row, 1):
        data_cell(ws1, ri, ci, val, alt)

ws1.freeze_panes = "A2"

# ── Sheet 2: Care Gaps ────────────────────────────────────────────────────────
ws2 = wb.create_sheet("Care Gap File")
ws2.row_dimensions[1].height = 28

GAP_COLS = [
    ("member_id", 14), ("measure_code", 14), ("plan_id", 10),
    ("measurement_year", 16), ("gap_status", 12), ("gap_open_date", 14),
    ("clinical_risk_score", 18), ("days_open", 11),
    ("nba_propensity_score", 18), ("upstream_recommended_channel", 24),
    ("upstream_recommended_incentive", 26),
]
for ci, (name, w) in enumerate(GAP_COLS, 1):
    hdr_cell(ws2, 1, ci, name, w)

for ri, row in enumerate(gaps, 2):
    alt = ri % 2 == 0
    for ci, val in enumerate(row, 1):
        c = data_cell(ws2, ri, ci, val, alt)
        if ci in (7, 9):  # scores
            c.number_format = "0.00"

ws2.freeze_panes = "A2"

# ── Sheet 3: Plans ────────────────────────────────────────────────────────────
ws3 = wb.create_sheet("Plan Configuration")
ws3.row_dimensions[1].height = 28

PLAN_COLS = [
    ("plan_id", 10), ("plan_name", 36), ("region", 12), ("segment", 10),
    ("current_star_rating", 18), ("target_star_rating", 16),
    ("annual_revenue", 18), ("total_members", 14),
]
for ci, (name, w) in enumerate(PLAN_COLS, 1):
    hdr_cell(ws3, 1, ci, name, w)

for ri, row in enumerate(PLANS, 2):
    alt = ri % 2 == 0
    for ci, val in enumerate(row, 1):
        c = data_cell(ws3, ri, ci, val, alt)
        if ci in (5, 6):  # star ratings
            c.number_format = "0.0"
        if ci == 7:       # revenue
            c.number_format = "#,##0"
            c.alignment = Alignment(horizontal="right", vertical="center")
        if ci == 8:       # members
            c.number_format = "#,##0"
            c.alignment = Alignment(horizontal="right", vertical="center")

ws3.freeze_panes = "A2"

# ── Instructions sheet ────────────────────────────────────────────────────────
ws4 = wb.create_sheet("Instructions", 0)
ws4.sheet_properties.tabColor = ORANGE
instructions = [
    ("CareIntel — Demo Data Package", True, 16),
    ("", False, 11),
    ("This workbook contains sample data for testing the CareIntel Data Sources tab.", False, 11),
    ("Upload each sheet as a separate file using the correct data type:", False, 11),
    ("", False, 11),
    ("  1.  Member Roster  →  upload as 'Member Roster' data type", False, 11),
    ("  2.  Care Gap File  →  upload as 'Care Gap File' data type", False, 11),
    ("  3.  Plan Configuration  →  upload as 'Plan Configuration' data type", False, 11),
    ("", False, 11),
    ("Star Ratings in this dataset:", True, 11),
    ("  • Aetna Medicare Choice PPO (Northeast)    3.5 → 4.0", False, 11),
    ("  • Aetna Medicare Premier PPO (Southeast)   4.0 → 4.5", False, 11),
    ("  • Aetna Medicare DSNP Community (Midwest)  3.0 → 4.0", False, 11),
    ("  • UHC Medicare Advantage Value (West)       4.5 → 5.0", False, 11),
    ("  • UHC Medicare Signature PPO (West)         2.5 → 3.0", False, 11),
    ("", False, 11),
    (f"Generated: {date.today().strftime('%B %d, %Y')}  |  {len(members)} members  |  {len(gaps)} care gaps", False, 10),
]
ws4.column_dimensions["A"].width = 70
for ri, (text, bold, sz) in enumerate(instructions, 2):
    c = ws4.cell(row=ri, column=1, value=text)
    c.font = Font(bold=bold, size=sz, color="2D2D2D" if not bold else ORANGE)
    c.alignment = Alignment(vertical="center")
    ws4.row_dimensions[ri].height = 20 if text else 8

out = os.path.join(os.path.dirname(__file__), "careintel_demo_data.xlsx")
wb.save(out)
print(f"Saved: {out}")
print(f"  Members: {len(members)}")
print(f"  Gaps:    {len(gaps)}")
print(f"  Plans:   {len(PLANS)}")
