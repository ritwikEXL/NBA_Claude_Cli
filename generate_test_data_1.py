"""
Generate test_data_1.xlsx — Medicare Stars NBA demo Excel file
with fictional Medicare Advantage company data.
"""

import random
from datetime import date, timedelta
from openpyxl import Workbook

random.seed(42)

# ── Plan Configuration ────────────────────────────────────────────────────────

PLANS = [
    {"plan_id": "P101", "plan_name": "Humana Gold Plus HMO",         "region": "Southeast",   "segment": "HMO",  "current_star_rating": 3.0, "target_star_rating": 3.5, "annual_revenue": 520_000_000, "total_members": 48_000},
    {"plan_id": "P102", "plan_name": "Cigna HealthSpring Preferred",  "region": "Midwest",     "segment": "PPO",  "current_star_rating": 3.5, "target_star_rating": 4.0, "annual_revenue": 680_000_000, "total_members": 63_000},
    {"plan_id": "P103", "plan_name": "Molina Healthcare Select",      "region": "Southwest",   "segment": "HMO",  "current_star_rating": 2.5, "target_star_rating": 3.5, "annual_revenue": 310_000_000, "total_members": 28_500},
    {"plan_id": "P104", "plan_name": "WellCare Advance",              "region": "South",       "segment": "HMO",  "current_star_rating": 3.0, "target_star_rating": 4.0, "annual_revenue": 410_000_000, "total_members": 38_000},
    {"plan_id": "P105", "plan_name": "BCBS of Tennessee BlueMedicare","region": "Mid-Atlantic", "segment": "PPO",  "current_star_rating": 4.0, "target_star_rating": 4.5, "annual_revenue": 890_000_000, "total_members": 82_000},
    {"plan_id": "P106", "plan_name": "Highmark Freedom Blue PPO",     "region": "Northeast",   "segment": "PPO",  "current_star_rating": 3.5, "target_star_rating": 4.5, "annual_revenue": 750_000_000, "total_members": 69_000},
]

PLAN_IDS = [p["plan_id"] for p in PLANS]

# ── Member Roster ─────────────────────────────────────────────────────────────

LANGUAGES       = ["English", "Spanish", "Mandarin", "Vietnamese", "Korean", "Tagalog", "English"]
DIGITAL_SEGS    = ["High", "Medium", "Low", "None"]
DIGITAL_WEIGHTS = [0.25, 0.35, 0.25, 0.15]
SES_SEGS        = ["Urban-Middle", "Urban-Low", "Suburban-Middle", "Rural-Low", "Rural-Middle"]
CHANNELS        = ["Email", "SMS", "Phone", "Mail"]
GENDERS         = ["M", "F", "F", "M", "F"]   # slight female skew (realistic 65+)

def random_dob(seed_member_id: int) -> str:
    """Return a date_of_birth string for an age 65-90."""
    age_days = random.randint(65 * 365, 90 * 365)
    dob = date(2026, 7, 29) - timedelta(days=age_days)
    return dob.isoformat()

def channel_perms(pref: str):
    """Return (email_permission, sms_permission, call_permission, do_not_contact)."""
    dnc = random.random() < 0.05
    if dnc:
        return "N", "N", "N", "Y"
    email = "Y" if pref in ("Email",)     or random.random() < 0.6 else "N"
    sms   = "Y" if pref in ("SMS",)       or random.random() < 0.5 else "N"
    call  = "Y" if pref in ("Phone",)     or random.random() < 0.7 else "N"
    return email, sms, call, "N"

members = []
for i in range(1, 151):
    mid   = f"M{1000 + i}"
    pref  = random.choice(CHANNELS)
    lang  = random.choices(LANGUAGES, weights=[50, 20, 8, 6, 5, 5, 6])[0]
    dig   = random.choices(DIGITAL_SEGS, weights=DIGITAL_WEIGHTS)[0]
    ses   = random.choice(SES_SEGS)
    dob   = random_dob(i)
    gen   = random.choice(GENDERS)
    ep, sp, cp, dnc = channel_perms(pref)
    notes = ""
    if lang != "English":
        notes = f"Prefers {lang}-language materials"
    if dig == "None":
        notes = (notes + "; No digital access").lstrip("; ")
    members.append({
        "member_id": mid,
        "date_of_birth": dob,
        "gender": gen,
        "language_preference": lang,
        "digital_literacy_segment": dig,
        "socioeconomic_segment": ses,
        "preferred_channel": pref,
        "email_permission": ep,
        "sms_permission": sp,
        "call_permission": cp,
        "do_not_contact": dnc,
        "channel_preference_notes": notes,
    })

# ── Care Gap File ─────────────────────────────────────────────────────────────

MEASURES = ["BCS", "COL", "EED", "CBP"]
MEASURE_WEIGHTS = [0.30, 0.25, 0.25, 0.20]

# Age eligibility filters (approximate)
MEASURE_AGE = {
    "BCS": (50, 74),   # female only ideally, but we keep it simple
    "COL": (45, 75),
    "EED": (18, 90),
    "CBP": (18, 85),
}

NBA_RECS = {
    "BCS": [
        "Schedule mammogram via telehealth referral",
        "Send SMS reminder with in-network radiology link",
        "Outreach call — offer transport voucher for screening",
        "Mail reminder with pre-scheduled appointment slip",
    ],
    "COL": [
        "Mail FIT kit with prepaid return envelope",
        "SMS alert: colorectal screening overdue, easy home test available",
        "Outreach call — offer colonoscopy scheduling assistance",
        "Email with telehealth gastroenterology referral link",
    ],
    "EED": [
        "SMS reminder: annual diabetic eye exam overdue",
        "Outreach call — schedule retinal screening at nearest clinic",
        "Mail reminder with in-network ophthalmologist list",
        "Email: complete eye exam to protect your vision — schedule now",
    ],
    "CBP": [
        "SMS: blood pressure check reminder — visit any CVS MinuteClinic",
        "Outreach call — in-home BP monitoring kit available",
        "Email: manage your blood pressure — schedule telehealth visit",
        "Mail: blood pressure tracking log and local resource guide",
    ],
}

gaps = []
used_combos = set()  # (member_id, measure_code) — one per member per measure

# Assign each member 1-3 gaps
for m in members:
    n_gaps = random.choices([1, 2, 3], weights=[0.35, 0.45, 0.20])[0]
    measures_for_member = random.choices(MEASURES, weights=MEASURE_WEIGHTS, k=n_gaps * 3)
    seen_measures = set()
    assigned = 0
    for msr in measures_for_member:
        if assigned >= n_gaps:
            break
        combo = (m["member_id"], msr)
        if combo in used_combos:
            continue
        used_combos.add(combo)
        seen_measures.add(msr)
        assigned += 1

        plan_id      = random.choice(PLAN_IDS)
        meas_year    = random.choice([2024, 2025])
        gap_status   = random.choices(["open", "closed"], weights=[0.70, 0.30])[0]
        days_open    = random.randint(30, 400) if gap_status == "open" else random.randint(1, 200)
        gap_open_dt  = (date(2026, 7, 29) - timedelta(days=days_open)).isoformat()
        risk_score   = round(random.uniform(0.30, 0.95), 4)
        nba_rec      = random.choice(NBA_RECS[msr])

        gaps.append({
            "member_id":           m["member_id"],
            "measure_code":        msr,
            "plan_id":             plan_id,
            "measurement_year":    meas_year,
            "gap_status":          gap_status,
            "gap_open_date":       gap_open_dt,
            "clinical_risk_score": risk_score,
            "days_open":           days_open,
            "nba_recommendation":  nba_rec,
        })

# Pad to 250+ if needed
while len(gaps) < 250:
    m = random.choice(members)
    msr = random.choice(MEASURES)
    combo = (m["member_id"], msr)
    if combo in used_combos:
        continue
    used_combos.add(combo)
    plan_id   = random.choice(PLAN_IDS)
    meas_year = random.choice([2024, 2025])
    gap_status = random.choices(["open", "closed"], weights=[0.70, 0.30])[0]
    days_open  = random.randint(30, 400) if gap_status == "open" else random.randint(1, 200)
    gap_open_dt = (date(2026, 7, 29) - timedelta(days=days_open)).isoformat()
    risk_score  = round(random.uniform(0.30, 0.95), 4)
    nba_rec     = random.choice(NBA_RECS[msr])
    gaps.append({
        "member_id":           m["member_id"],
        "measure_code":        msr,
        "plan_id":             plan_id,
        "measurement_year":    meas_year,
        "gap_status":          gap_status,
        "gap_open_date":       gap_open_dt,
        "clinical_risk_score": risk_score,
        "days_open":           days_open,
        "nba_recommendation":  nba_rec,
    })

# ── Write Excel ───────────────────────────────────────────────────────────────

OUTPUT_PATH = r"C:\Users\vmuser\Documents\NBA_Claude_Cli\test_data_1.xlsx"

wb = Workbook()

# Sheet 1 — Plan Configuration
ws1 = wb.active
ws1.title = "Plan Configuration"
plan_cols = ["plan_id", "plan_name", "region", "segment",
             "current_star_rating", "target_star_rating", "annual_revenue", "total_members"]
ws1.append(plan_cols)
for p in PLANS:
    ws1.append([p[c] for c in plan_cols])

# Sheet 2 — Member Roster
ws2 = wb.create_sheet("Member Roster")
member_cols = [
    "member_id", "date_of_birth", "gender", "language_preference",
    "digital_literacy_segment", "socioeconomic_segment", "preferred_channel",
    "email_permission", "sms_permission", "call_permission",
    "do_not_contact", "channel_preference_notes",
]
ws2.append(member_cols)
for m in members:
    ws2.append([m[c] for c in member_cols])

# Sheet 3 — Care Gap File
ws3 = wb.create_sheet("Care Gap File")
gap_cols = [
    "member_id", "measure_code", "plan_id", "measurement_year",
    "gap_status", "gap_open_date", "clinical_risk_score",
    "days_open", "nba_recommendation",
]
ws3.append(gap_cols)
for g in gaps:
    ws3.append([g[c] for c in gap_cols])

wb.save(OUTPUT_PATH)

print(f"File written: {OUTPUT_PATH}")
print(f"  Sheet 'Plan Configuration' : {len(PLANS)} data rows")
print(f"  Sheet 'Member Roster'      : {len(members)} data rows")
print(f"  Sheet 'Care Gap File'      : {len(gaps)} data rows")
