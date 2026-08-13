"""Generate test_data_3.xlsx — 6 plans, ~142 members, ~559 gaps. call_allowed guaranteed."""
import random
from datetime import date, timedelta
from openpyxl import Workbook

random.seed(99)

PLANS = [
    ("P301", "Bright Health Medicare",              "Southeast",   "MAPD", 2.5, 4.0,  90_000_000,  8_300),
    ("P302", "Molina Healthcare Dual Plus",         "Southwest",   "MAPD", 3.0, 4.0, 195_000_000, 18_000),
    ("P303", "Zing Health Medicare Advantage",      "Midwest",     "MAPD", 2.0, 3.5,  55_000_000,  5_100),
    ("P304", "Alignment Health Plan",               "West",        "MAPD", 3.5, 4.5, 140_000_000, 12_900),
    ("P305", "Friday Health Plans Medicare",        "South",       "MAPD", 2.5, 3.5,  75_000_000,  6_900),
    ("P306", "Innovage PACE",                       "Northeast",   "PACE", 3.0, 4.0,  48_000_000,  4_400),
]

MEASURES = ["BCS", "COL", "EED", "CBP", "MAD"]

FIRST_NAMES = ["Alice","Barbara","Carol","Diana","Eleanor","Frances","Grace","Helen",
               "Irene","Janet","Karen","Linda","Margaret","Nancy","Olivia","Patricia",
               "Rachel","Sandra","Teresa","Ursula","Victoria","Wendy","Yvonne",
               "James","Robert","John","Michael","William","David","Richard","Joseph",
               "Thomas","Charles","Gary","Donald","Kenneth","Steven","Edward","Brian",
               "Ronald","Anthony","Kevin","Jason","Matthew","Frank","George","Harold",
               "Larry","Douglas","Arthur","Henry","Walter","Dennis","Jerry","Wayne"]

LAST_NAMES = ["Smith","Johnson","Williams","Brown","Jones","Garcia","Martinez","Davis",
              "Lopez","Wilson","Anderson","Taylor","Moore","Jackson","Martin","Lee",
              "Thompson","White","Harris","Clark","Lewis","Robinson","Walker","Hall",
              "Young","Allen","King","Wright","Scott","Green","Baker","Adams","Nelson",
              "Carter","Mitchell","Perez","Roberts","Turner","Phillips","Campbell"]

LANGUAGES = ["English"]*70 + ["Spanish"]*12 + ["Mandarin"]*5 + ["Vietnamese"]*5 + \
            ["Korean"]*4 + ["Tagalog"]*4
LITERACY  = ["High"]*40 + ["Medium"]*40 + ["Low"]*20

def rand_dob():
    start = date(1940, 1, 1)
    end   = date(1965, 12, 31)
    return start + timedelta(days=random.randint(0, (end - start).days))

# ── Members ───────────────────────────────────────────────────────────────────
members = []
mid = 3001
for plan in PLANS:
    count = random.randint(21, 26)
    for _ in range(count):
        lang = random.choice(LANGUAGES)
        lit  = random.choice(LITERACY)
        email = 1 if random.random() < 0.55 else 0
        sms   = 1 if (random.random() < 0.45 if lang == "English" else random.random() < 0.25) else 0
        call  = 1  # guaranteed — no suppression in test_data_3
        if email:   pref = "Email"
        elif sms:   pref = "SMS"
        else:       pref = "Phone"
        members.append({
            "member_id":        f"M{mid}",
            "plan_id":          plan[0],
            "first_name":       random.choice(FIRST_NAMES),
            "last_name":        random.choice(LAST_NAMES),
            "date_of_birth":    rand_dob().isoformat(),
            "gender":           random.choice(["F","F","F","M","M"]),
            "language":         lang,
            "digital_literacy": lit,
            "email_allowed":    email,
            "sms_allowed":      sms,
            "call_allowed":     call,
            "preferred_channel":pref,
        })
        mid += 1

print(f"Members: {len(members)}")

# ── Gaps ──────────────────────────────────────────────────────────────────────
gaps = []
for m in members:
    plan_star = next(p[4] for p in PLANS if p[0] == m["plan_id"])
    open_prob = 0.75 if plan_star < 3.0 else (0.60 if plan_star < 3.5 else 0.45)

    num_measures = random.randint(3, 5)
    chosen = random.sample(MEASURES, num_measures)
    for mcode in chosen:
        roll = random.random()
        if roll < open_prob:
            status   = "open"
            days_open = random.randint(30, 400)
        elif roll < open_prob + 0.10:
            status   = "borderline"
            days_open = random.randint(10, 90)
        else:
            status   = "closed"
            days_open = 0

        base = {"High": 0.65, "Medium": 0.50, "Low": 0.35}[m["digital_literacy"]]
        prop = round(min(0.95, max(0.05, base + random.gauss(0, 0.15))), 2)

        gaps.append({
            "member_id":            m["member_id"],
            "plan_id":              m["plan_id"],
            "measure_code":         mcode,
            "gap_status":           status,
            "nba_propensity_score": prop,
            "days_open":            days_open,
        })

print(f"Gaps: {len(gaps)}")

# ── Write workbook ────────────────────────────────────────────────────────────
wb = Workbook()

ws1 = wb.active
ws1.title = "Plan Configuration"
ws1.append(["plan_id","plan_name","region","segment","current_star_rating","target_star_rating","annual_revenue","total_members"])
for p in PLANS:
    ws1.append(list(p))

ws2 = wb.create_sheet("Member Roster")
cols2 = ["member_id","plan_id","first_name","last_name","date_of_birth","gender",
         "language","digital_literacy","email_allowed","sms_allowed","call_allowed","preferred_channel"]
ws2.append(cols2)
for m in members:
    ws2.append([m[c] for c in cols2])

ws3 = wb.create_sheet("Care Gap File")
cols3 = ["member_id","plan_id","measure_code","gap_status","nba_propensity_score","days_open"]
ws3.append(cols3)
for g in gaps:
    ws3.append([g[c] for c in cols3])

out = r"C:\Users\vmuser\Documents\NBA_Claude_Cli\test_data_3.xlsx"
wb.save(out)
print(f"Saved: {out}")
print(f"Plans: {len(PLANS)}, Members: {len(members)}, Gaps: {len(gaps)}")
